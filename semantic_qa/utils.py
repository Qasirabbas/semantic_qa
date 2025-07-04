import re
import pandas as pd
import openpyxl
from typing import List, Dict, Optional
from django.core.files.uploadedfile import UploadedFile
from io import BytesIO
import logging

logger = logging.getLogger('semantic_qa')

def get_client_ip(request):
    """Get client IP address from request"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def clean_text(text: str) -> str:
    """Clean and normalize text"""
    if not text:
        return ""
    
    # Remove extra whitespace and normalize
    text = re.sub(r'\s+', ' ', str(text)).strip()
    
    # Remove special characters but keep basic punctuation
    #modify history
    # roar edit 2025-6-17 
    # text = re.sub(r'[^\w\s\-_.,!?()]', '', text)
    
    return text

def extract_keywords_from_text(text: str) -> str:
    """Extract keywords from question and answer for better searching - supports Chinese"""
    if not text:
        return ""
    
    # Detect if text contains Chinese characters
    def contains_chinese(text):
        return any('\u4e00' <= char <= '\u9fff' for char in text)
    
    has_chinese = contains_chinese(text)
    
    if has_chinese:
        # Chinese text processing
        # Extract Chinese words (sequences of Chinese characters)
        chinese_words = re.findall(r'[\u4e00-\u9fff]+', text)
        
        # Extract English words and numbers
        english_words = re.findall(r'[a-zA-Z0-9]+', text)
        
        # Common Chinese stop words
        chinese_stop_words = {
            '的', '了', '在', '是', '我', '有', '和', '就', '不', '人', '都', '一', '一个', 
            '上', '也', '很', '到', '说', '要', '去', '你', '会', '着', '没有', '看', '好',
            '自己', '这', '那', '什么', '怎么', '如何', '为什么', '哪里', '什么时候',
            '怎样', '为了', '因为', '所以', '但是', '然后', '现在', '已经', '可以',
            '应该', '需要', '必须', '能够', '或者', '还是', '如果', '虽然', '问', '答'
        }
        
        # Filter Chinese words
        filtered_chinese = [word for word in chinese_words 
                           if word not in chinese_stop_words and len(word) > 1]
        
        # Common English stop words
        english_stop_words = {
            'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 
            'of', 'with', 'by', 'how', 'what', 'where', 'when', 'why', 'is', 'are', 
            'was', 'were', 'be', 'been', 'being', 'have', 'has', 'had', 'do', 'does', 
            'did', 'will', 'would', 'could', 'should', 'can', 'may', 'might', 'must',
            'this', 'that', 'these', 'those', 'i', 'you', 'he', 'she', 'it', 'we', 'they'
        }
        
        # Filter English words
        filtered_english = [word.lower() for word in english_words 
                           if word.lower() not in english_stop_words and len(word) > 2]
        
        # Combine all keywords
        all_keywords = filtered_chinese + filtered_english
        
        # Remove duplicates while preserving order
        seen = set()
        unique_keywords = []
        for keyword in all_keywords:
            if keyword not in seen:
                seen.add(keyword)
                unique_keywords.append(keyword)
        
        return ' '.join(unique_keywords[:20])  # Limit to 20 keywords
    
    else:
        # English text processing (original logic)
        text = text.lower()
        
        # Remove common stop words
        stop_words = {
            'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 
            'of', 'with', 'by', 'how', 'what', 'where', 'when', 'why', 'is', 'are', 
            'was', 'were', 'be', 'been', 'being', 'have', 'has', 'had', 'do', 'does', 
            'did', 'will', 'would', 'could', 'should', 'can', 'may', 'might', 'must',
            'this', 'that', 'these', 'those', 'i', 'you', 'he', 'she', 'it', 'we', 'they'
        }
        
        # Extract words and filter
        words = re.findall(r'\b\w+\b', text)
        keywords = [word for word in words if word not in stop_words and len(word) > 2]
        
        # Remove duplicates while preserving order
        seen = set()
        unique_keywords = []
        for keyword in keywords:
            if keyword not in seen:
                seen.add(keyword)
                unique_keywords.append(keyword)
        
        return ' '.join(unique_keywords[:15])  # Limit to 15 keywords for English
def clean_image_url(url: str) -> str:
    """Clean and fix image URLs - ENHANCED VERSION"""
    if not url:
        return ""
    
    url = url.strip()
    
    # Fix common URL formatting issues
    
    # Case 1: Missing :// after protocol
    if 'httpsae01' in url:
        url = url.replace('httpsae01', 'https://ae01')
    elif 'httpae01' in url:
        url = url.replace('httpae01', 'http://ae01')
    elif 'httpswww' in url:
        url = url.replace('httpswww', 'https://www')
    elif 'httpwww' in url:
        url = url.replace('httpwww', 'http://www')
    
    # Case 2: Missing slash after domain - SPECIFIC FOR ALICDN
    if 'ae01.alicdn.comkf' in url:
        url = url.replace('ae01.alicdn.comkf', 'ae01.alicdn.com/kf/')
    
    # Case 3: General missing :// fix
    if url.startswith('https') and '://' not in url:
        url = url.replace('https', 'https://', 1)
    elif url.startswith('http') and '://' not in url:
        url = url.replace('http', 'http://', 1)
    
    # Case 4: URL doesn't start with protocol at all
    elif not url.startswith(('http://', 'https://')):
        if 'alicdn.com' in url or 'ae01' in url:
            url = 'https://' + url
        else:
            url = 'https://' + url
    
    return url

def parse_excel_file(file: UploadedFile) -> List[Dict]:
    """Parse Excel file and extract QA entries with multilingual support"""
    try:
        # Read Excel file
        if file.name.endswith('.xlsx'):
            df = pd.read_excel(file, engine='openpyxl')
        else:
            df = pd.read_excel(file, engine='xlrd')
        
        # Clean column names - remove extra spaces and normalize
        df.columns = df.columns.str.strip()
        
        # Print columns for debugging
        logger.info(f"Excel columns found: {list(df.columns)}")
        
        # Define multilingual column mappings
        column_mappings = {
            'SKU': {
                'en': ['sku', 'product_id', 'part_number', 'product id', 'part number', 'code', 'id'],
                'zh': ['sku', '产品编号', '产品代码', '编号', '代码', '产品号', '型号', '货号'],
                'variations': ['sku', 'product_id', 'part_number', 'productid', 'partnumber']
            },
            'Question': {
                'en': ['question', 'q', 'query', 'problem', 'issue', 'ask'],
                'zh': ['问题', '疑问', '询问', '内容', '题目', 'q', 'question'],
                'variations': ['question', 'q', 'query', 'problem']
            },
            'Answer': {
                'en': ['answer', 'a', 'response', 'reply', 'solution', 'result'],
                'zh': ['答案', '回答', '解答', '解决方案','内容', '回复', 'a', 'answer'],
                'variations': ['answer', 'a', 'response', 'reply']
            },
            'Image_Link': {
                'en': ['image_link', 'image', 'img_link', 'image_url', 'imagelink', 'picture', 'photo'],
                'zh': ['图片链接', '图片', '照片', '图像', '链接', '图片地址', '图片网址'],
                'variations': ['image_link', 'image', 'img_link', 'imagelink']
            },
            'Category': {
                'en': ['category', 'cat', 'type', 'class', 'group'],
                'zh': ['类别', '分类', '类型', '种类', '组别', '类'],
                'variations': ['category', 'cat', 'type', 'class']
            },
            'Keywords': {
                'en': ['keywords', 'keyword', 'tags', 'search_terms', 'search terms'],
                'zh': ['关键词', '关键字', '标签', '搜索词', '检索词'],
                'variations': ['keywords', 'keyword', 'tags']
            }
        }
        
        # Expected columns
        required_columns = ['SKU', 'Question', 'Answer']
        
        # Create column mapping
        column_mapping = {}
        
        # Convert all column names to lowercase for comparison
        df_columns_lower = {col.lower(): col for col in df.columns}
        
        # Also create mapping without spaces/underscores
        df_columns_normalized = {}
        for col in df.columns:
            normalized = col.lower().replace(' ', '').replace('_', '').replace('-', '')
            df_columns_normalized[normalized] = col
        
        # Map required columns with multilingual support
        for req_col in required_columns:
            found = False
            mapping_config = column_mappings.get(req_col, {})
            
            # Collect all possible names from all languages
            all_possible_names = []
            
            # Add English variations
            all_possible_names.extend(mapping_config.get('en', []))
            # Add Chinese variations  
            all_possible_names.extend(mapping_config.get('zh', []))
            # Add common variations
            all_possible_names.extend(mapping_config.get('variations', []))
            # Add the original column name
            all_possible_names.append(req_col.lower())
            
            # Try exact matching first
            for possible_name in all_possible_names:
                possible_name_lower = possible_name.lower()
                if possible_name_lower in df_columns_lower:
                    column_mapping[req_col] = df_columns_lower[possible_name_lower]
                    found = True
                    logger.info(f"Mapped {req_col} to '{df_columns_lower[possible_name_lower]}' via exact match")
                    break
            
            # Try normalized matching (without spaces/underscores)
            if not found:
                for possible_name in all_possible_names:
                    normalized_name = possible_name.lower().replace(' ', '').replace('_', '').replace('-', '')
                    if normalized_name in df_columns_normalized:
                        column_mapping[req_col] = df_columns_normalized[normalized_name]
                        found = True
                        logger.info(f"Mapped {req_col} to '{df_columns_normalized[normalized_name]}' via normalized match")
                        break
            
            # Try partial matching
            if not found:
                for col_lower, original_col in df_columns_lower.items():
                    for possible_name in all_possible_names:
                        possible_name_lower = possible_name.lower()
                        if (possible_name_lower in col_lower or col_lower in possible_name_lower) and len(possible_name_lower) > 2:
                            column_mapping[req_col] = original_col
                            found = True
                            logger.info(f"Mapped {req_col} to '{original_col}' via partial match with '{possible_name}'")
                            break
                    if found:
                        break
            
            # Special handling for Chinese files - try positional mapping as last resort
            if not found and len(df.columns) >= len(required_columns):
                # For Chinese files, common order is: Index, Content/Question, Image
                # Try to map based on position if we can identify the pattern
                if req_col == 'Question' and len(df.columns) >= 2:
                    # Usually the second column (index 1) contains the main content
                    if '内容' in df.columns[1] or '问题' in df.columns[1] or not df.columns[1].startswith('Unnamed'):
                        column_mapping[req_col] = df.columns[1]
                        found = True
                        logger.info(f"Mapped {req_col} to '{df.columns[1]}' via positional mapping")
                elif req_col == 'SKU' and len(df.columns) >= 1:
                    # If first column has data (not just Unnamed), use it as SKU
                    if not df.columns[0].startswith('Unnamed') or df.iloc[0, 0] is not None:
                        column_mapping[req_col] = df.columns[0]
                        found = True
                        logger.info(f"Mapped {req_col} to '{df.columns[0]}' via positional mapping")
                elif req_col == 'Answer':
                    # For Answer, we'll generate it from Question if not found
                    pass
            
            if not found and req_col != 'Answer':
                # List available columns for debugging
                available_cols = list(df.columns)
                logger.error(f"Required column '{req_col}' not found. Available columns: {available_cols}")
                
                # Provide helpful suggestions based on what we found
                suggestions = []
                if any('内容' in col for col in df.columns):
                    suggestions.append("Found Chinese content column - this might be your Question column")
                if any('图片' in col or 'image' in col.lower() for col in df.columns):
                    suggestions.append("Found image-related column")
                    
                suggestion_text = ". Suggestions: " + "; ".join(suggestions) if suggestions else ""
                
                raise ValueError(f"Required column '{req_col}' not found in Excel file. Available columns: {available_cols}{suggestion_text}")
        
        # Map optional columns with the same multilingual approach
        optional_columns = ['Image_Link', 'Category', 'Keywords']
        for opt_col in optional_columns:
            mapping_config = column_mappings.get(opt_col, {})
            
            # Collect all possible names
            all_possible_names = []
            all_possible_names.extend(mapping_config.get('en', []))
            all_possible_names.extend(mapping_config.get('zh', []))
            all_possible_names.extend(mapping_config.get('variations', []))
            all_possible_names.append(opt_col.lower())
            
            # Try to find the column
            for possible_name in all_possible_names:
                possible_name_lower = possible_name.lower()
                if possible_name_lower in df_columns_lower:
                    column_mapping[opt_col] = df_columns_lower[possible_name_lower]
                    logger.info(f"Mapped optional {opt_col} to '{df_columns_lower[possible_name_lower]}'")
                    break
        
        logger.info(f"Final column mapping: {column_mapping}")
        
        qa_entries = []
        
        # Process rows
        for index, row in df.iterrows():
            try:
                # Get SKU
                sku = ""
                if 'SKU' in column_mapping:
                    sku_value = row[column_mapping['SKU']]
                    if pd.notna(sku_value):
                        sku = clean_text(str(sku_value))
                
                # If no SKU found, generate one from row index
                if not sku:
                    sku = f"ITEM_{index + 1:03d}"
                
                # Get Question
                question = ""
                if 'Question' in column_mapping:
                    question_value = row[column_mapping['Question']]
                    if pd.notna(question_value):
                        question = clean_text(str(question_value))
                
                # Skip if no question content
                if not question:
                    logger.warning(f"Skipping row {index + 1}: no question content found")
                    continue
                
                # Get Answer - if not found, use question as answer
                answer = ""
                if 'Answer' in column_mapping:
                    answer_value = row[column_mapping['Answer']]
                    if pd.notna(answer_value):
                        answer = clean_text(str(answer_value))
                
                # If no answer, use question as answer (common in FAQ format)
                if not answer:
                    answer = question
                    logger.info(f"Row {index + 1}: Using question as answer (FAQ format)")
                
                # Extract optional fields
                image_link = ""
                if 'Image_Link' in column_mapping:
                    image_value = row[column_mapping['Image_Link']]
                    if pd.notna(image_value):
                        image_link = clean_text(str(image_value))
                        image_link = clean_image_url(image_link)
                
                category = ""
                if 'Category' in column_mapping:
                    cat_value = row[column_mapping['Category']]
                    if pd.notna(cat_value):
                        category = clean_text(str(cat_value))
                
                keywords = ""
                if 'Keywords' in column_mapping:
                    kw_value = row[column_mapping['Keywords']]
                    if pd.notna(kw_value):
                        keywords = clean_text(str(kw_value))
                
                # If no keywords provided, extract from question and answer
                if not keywords:
                    keywords = extract_keywords_from_text(f"{question} {answer}")
                
                # Determine category if not provided
                if not category:
                    category = determine_category(question, answer)
                
                qa_entry = {
                    'sku': sku,
                    'question': question,
                    'answer': answer,
                    'image_link': image_link,
                    'category': category,
                    'keywords': keywords
                }
                
                qa_entries.append(qa_entry)
                logger.debug(f"Processed row {index + 1}: SKU={sku}, Question={question[:50]}...")
                
            except Exception as e:
                logger.error(f"Error processing row {index + 1}: {str(e)}")
                continue
        
        logger.info(f"Successfully parsed {len(qa_entries)} entries from Excel file")
        return qa_entries
        
    except Exception as e:
        logger.error(f"Error parsing Excel file: {str(e)}")
        raise ValueError(f"Error parsing Excel file: {str(e)}")

# def extract_keywords_from_text(text: str) -> str:
#     """Extract keywords from question and answer for better searching - supports Chinese"""
#     if not text:
#         return ""
    
#     # Detect if text contains Chinese characters
#     def contains_chinese(text):
#         return any('\u4e00' <= char <= '\u9fff' for char in text)
    
#     has_chinese = contains_chinese(text)
    
#     if has_chinese:
#         # Chinese text processing
#         # For Chinese, we extract meaningful phrases and words
#         import re
        
#         # Extract Chinese words (sequences of Chinese characters)
#         chinese_words = re.findall(r'[\u4e00-\u9fff]+', text)
        
#         # Extract English words and numbers
#         english_words = re.findall(r'[a-zA-Z0-9]+', text)
        
#         # Common Chinese stop words
#         chinese_stop_words = {
#             '的', '了', '在', '是', '我', '有', '和', '就', '不', '人', '都', '一', '一个', 
#             '上', '也', '很', '到', '说', '要', '去', '你', '会', '着', '没有', '看', '好',
#             '自己', '这', '那', '什么', '怎么', '如何', '为什么', '哪里', '什么时候',
#             '怎样', '为了', '因为', '所以', '但是', '然后', '现在', '已经', '可以',
#             '应该', '需要', '必须', '能够', '或者', '还是', '如果', '虽然', '问', '答'
#         }
        
#         # Filter Chinese words
#         filtered_chinese = [word for word in chinese_words 
#                            if word not in chinese_stop_words and len(word) > 1]
        
#         # Common English stop words
#         english_stop_words = {
#             'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 
#             'of', 'with', 'by', 'how', 'what', 'where', 'when', 'why', 'is', 'are', 
#             'was', 'were', 'be', 'been', 'being', 'have', 'has', 'had', 'do', 'does', 
#             'did', 'will', 'would', 'could', 'should', 'can', 'may', 'might', 'must',
#             'this', 'that', 'these', 'those', 'i', 'you', 'he', 'she', 'it', 'we', 'they'
#         }
        
#         # Filter English words
#         filtered_english = [word.lower() for word in english_words 
#                            if word.lower() not in english_stop_words and len(word) > 2]
        
#         # Combine all keywords
#         all_keywords = filtered_chinese + filtered_english
        
#         # Remove duplicates while preserving order
#         seen = set()
#         unique_keywords = []
#         for keyword in all_keywords:
#             if keyword not in seen:
#                 seen.add(keyword)
#                 unique_keywords.append(keyword)
        
#         return ' '.join(unique_keywords[:20])  # Limit to 20 keywords
    
#     else:
#         # English text processing (original logic)
#         text = text.lower()
        
#         # Remove common stop words
#         stop_words = {
#             'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 
#             'of', 'with', 'by', 'how', 'what', 'where', 'when', 'why', 'is', 'are', 
#             'was', 'were', 'be', 'been', 'being', 'have', 'has', 'had', 'do', 'does', 
#             'did', 'will', 'would', 'could', 'should', 'can', 'may', 'might', 'must',
#             'this', 'that', 'these', 'those', 'i', 'you', 'he', 'she', 'it', 'we', 'they'
#         }
        
#         # Extract words and filter
#         words = re.findall(r'\b\w+\b', text)
#         keywords = [word for word in words if word not in stop_words and len(word) > 2]
        
#         # Remove duplicates while preserving order
#         seen = set()
#         unique_keywords = []
#         for keyword in keywords:
#             if keyword not in seen:
#                 seen.add(keyword)
#                 unique_keywords.append(keyword)
        
#         return ' '.join(unique_keywords[:15])  # Limit to 15 keywords for English


def determine_category(question: str, answer: str) -> str:
    """Automatically determine category based on question and answer content - supports Chinese"""
    text = f"{question} {answer}".lower()
    
    # Check if text contains Chinese
    def contains_chinese(text):
        return any('\u4e00' <= char <= '\u9fff' for char in text)
    
    if contains_chinese(text):
        # Chinese category detection
        chinese_category_keywords = {
            '安装': ['安装', '安设', '装配', '安装方法', '怎么装', '如何安装', '连接', '接线', '焊接'],
            '故障排除': ['故障', '问题', '不工作', '损坏', '修理', '维修', '排除', '检查', '不能', '无法'],
            '维护': ['维护', '保养', '清洁', '清理', '服务', '更换', '替换', '润滑', '定期'],
            '操作': ['操作', '使用', '怎么用', '如何使用', '控制', '功能', '运行'],
            '规格': ['规格', '尺寸', '大小', '重量', '材料', '容量', '参数', '技术参数'],
            '安全': ['安全', '危险', '警告', '注意', '防护', '保护', '安全性']
        }
        category_keywords = chinese_category_keywords
    else:
        # English category detection
        category_keywords = {
            'installation': ['install', 'installation', 'setup', 'mount', 'attach', 'connect', 'wiring', 'wire', 'cable', 'welding', 'weld'],
            'troubleshooting': ['problem', 'issue', 'error', 'not working', 'broken', 'fix', 'repair', 'troubleshoot', 'malfunction'],
            'maintenance': ['maintain', 'maintenance', 'clean', 'cleaning', 'service', 'replace', 'replacement', 'lubricate'],
            'operation': ['how to', 'operate', 'operation', 'use', 'usage', 'control', 'controls', 'function'],
            'specifications': ['spec', 'specification', 'dimension', 'size', 'weight', 'material', 'capacity'],
            'safety': ['safety', 'danger', 'warning', 'caution', 'hazard', 'protection', 'secure']
        }
    
    # Count keyword matches for each category
    category_scores = {}
    for category, keywords in category_keywords.items():
        score = sum(1 for keyword in keywords if keyword in text)
        if score > 0:
            category_scores[category] = score
    
    # Return category with highest score, or 'general' if no matches
    if category_scores:
        return max(category_scores, key=category_scores.get)
    else:
        return 'general'

def validate_image_url(url: str) -> bool:
    """Validate if URL is a valid image URL"""
    if not url:
        return False
    
    # Check if URL has image extension
    image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.svg']
    url_lower = url.lower()
    
    return any(url_lower.endswith(ext) for ext in image_extensions) or 'image' in url_lower

def sanitize_filename(filename: str) -> str:
    """Sanitize filename for safe storage"""
    # Remove or replace invalid characters
    filename = re.sub(r'[^\w\s\-_\.]', '', filename)
    filename = re.sub(r'\s+', '_', filename)
    return filename

# Add this function to your existing utils.py file

# Add this function to your existing utils.py file

def format_enhanced_search_results(results: List[Dict]) -> List[Dict]:
    """Enhanced format search results for display with proper score handling"""
    formatted_results = []
    
    for result in results:
        if result['type'] == 'qa_entry':
            entry = result['entry']
            
            # Determine relevance category and proper display for boosted scores
            score = result['score']
            
            # Enhanced display logic for boosted scores
            if score >= 1.0:  # Boosted scores (exact matches)
                relevance_category = 'exact'
                if score >= 1.3:
                    display_score = f"{score:.1f}★★★"  # Triple star for very high boost
                    relevance_text = "精确匹配"
                elif score >= 1.1:
                    display_score = f"{score:.1f}★★"   # Double star for high boost
                    relevance_text = "高度匹配"
                else:
                    display_score = f"{score:.1f}★"    # Single star for boost
                    relevance_text = "匹配加强"
            elif score >= 0.7:
                relevance_category = 'high'
                display_score = f"{score:.0%}"
                relevance_text = "高相关"
            elif score >= 0.4:
                relevance_category = 'medium'
                display_score = f"{score:.0%}"
                relevance_text = "中等相关"
            else:
                relevance_category = 'low'
                display_score = f"{score:.0%}"
                relevance_text = "可能相关"
            
            formatted_result = {
                'id': entry.id,
                'type': 'qa_entry',
                'sku': entry.sku,
                'question': entry.question,
                'answer': entry.answer,
                'image_link': entry.image_link,
                'category': entry.category or 'general',
                'relevance_score': score,
                'relevance_category': relevance_category,
                'relevance_text': relevance_text,
                'match_type': result['match_type'],
                'match_reason': result['match_reason'],
                'has_image': bool(entry.image_link),
                'display_score': display_score,
                'source_info': result.get('source_info', {}),
                'created_at': getattr(entry, 'created_at', None),
                'updated_at': getattr(entry, 'updated_at', None)
            }
            
            # Add formatted category for display
            if entry.category:
                formatted_result['category_display'] = entry.category.replace('_', ' ').title()
            else:
                formatted_result['category_display'] = '通用'
        
        elif result['type'] == 'document_chunk':
            chunk = result.get('chunk') or result.get('entry')  # Handle both formats
            document = result.get('document')
            
            # Determine relevance for document chunks
            score = result['score']
            if score >= 0.7:
                relevance_category = 'high'
                display_score = f"{score:.0%}"
                relevance_text = "高相关"
            elif score >= 0.4:
                relevance_category = 'medium'
                display_score = f"{score:.0%}"
                relevance_text = "中等相关"
            else:
                relevance_category = 'low'
                display_score = f"{score:.0%}"
                relevance_text = "可能相关"
            
            formatted_result = {
                'id': f"chunk_{chunk.id}",
                'type': 'document_chunk',
                'document_id': document.id if document else None,
                'document_title': document.title if document else 'Unknown Document',
                'document_type': document.document_type if document else 'unknown',
                'text_content': result.get('text_content', chunk.text if chunk else ''),
                'page_number': chunk.page_number if chunk else None,
                'chunk_index': chunk.chunk_index if chunk else None,
                'relevance_score': score,
                'relevance_category': relevance_category,
                'relevance_text': relevance_text,
                'match_type': result['match_type'],
                'match_reason': result['match_reason'],
                'display_score': display_score,
                'source_info': result.get('source_info', {}),
                'context_before': chunk.context_before if chunk else None,
                'context_after': chunk.context_after if chunk else None,
                'category': document.category if document else 'general',
                'category_display': '文档内容',
                'created_at': document.created_at if document else None
            }
        
        else:
            continue  # Skip unknown types
        
        # Add enhanced match type display
        match_type_display = {
            'exact_sku': '精确SKU匹配',
            'partial_sku': '部分SKU匹配', 
            'sku_prefix': 'SKU前缀匹配',
            'question_match': '问题匹配',
            'semantic': '语义匹配',
            'keyword': '关键词匹配'
        }
        
        formatted_result['match_type_display'] = match_type_display.get(
            result['match_type'], result['match_type']
        )
        
        formatted_results.append(formatted_result)
    
    return formatted_results

def generate_excel_template() -> BytesIO:
    """Generate multilingual Excel template for uploading QA data"""
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Create English template sheet
    ws_en = wb.active
    ws_en.title = "English_Template"
    
    # English headers
    headers_en = ['SKU', 'Question', 'Answer', 'Image_Link', 'Category', 'Keywords']
    
    # Add English headers with formatting
    for col, header in enumerate(headers_en, 1):
        cell = ws_en.cell(row=1, column=col, value=header)
        from openpyxl.styles import Font, PatternFill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Add English descriptions
    descriptions_en = [
        'Product identifier (required)',
        'Question text (required)', 
        'Answer text (required)',
        'Image URL (optional)',
        'Category name (optional)',
        'Search keywords (optional)'
    ]
    
    for col, desc in enumerate(descriptions_en, 1):
        ws_en.cell(row=2, column=col, value=desc)
    
    # Add English sample data
    sample_data_en = [
        ['ABC123', 'How to install the steering wheel?', 'Connect the wiring harness first, then secure with bolts.', 'https://example.com/steering_install.jpg', 'installation', 'steering wheel install wiring'],
        ['ABC123', 'Steering wheel controls not working?', 'Check the wiring connections behind the steering wheel.', 'https://example.com/steering_controls.jpg', 'troubleshooting', 'steering wheel controls troubleshooting'],
        ['XYZ789', 'How to weld the frame?', 'Use MIG welding with proper safety equipment.', 'https://example.com/welding.jpg', 'installation', 'welding frame MIG safety']
    ]
    
    for row, data in enumerate(sample_data_en, 3):
        for col, value in enumerate(data, 1):
            ws_en.cell(row=row, column=col, value=value)
    
    # Create Chinese template sheet
    ws_zh = wb.create_sheet("Chinese_Template")
    
    # Chinese headers
    headers_zh = ['SKU', '问题', '答案', '图片链接', '类别', '关键词']
    
    # Add Chinese headers with formatting
    for col, header in enumerate(headers_zh, 1):
        cell = ws_zh.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")
    
    # Add Chinese descriptions
    descriptions_zh = [
        '产品编号（必填）',
        '问题内容（必填）', 
        '答案内容（必填）',
        '图片网址（可选）',
        '分类名称（可选）',
        '搜索关键词（可选）'
    ]
    
    for col, desc in enumerate(descriptions_zh, 1):
        ws_zh.cell(row=2, column=col, value=desc)
    
    # Add Chinese sample data
    sample_data_zh = [
        ['ABC123', '如何安装方向盘？', '首先连接线束，然后用螺栓固定。', 'https://example.com/steering_install.jpg', '安装', '方向盘 安装 线束'],
        ['ABC123', '方向盘控制器不工作？', '检查方向盘后面的线路连接。', 'https://example.com/steering_controls.jpg', '故障排除', '方向盘 控制器 故障排除'],
        ['XYZ789', '如何焊接框架？', '使用MIG焊接并配备适当的安全设备。', 'https://example.com/welding.jpg', '安装', '焊接 框架 MIG 安全']
    ]
    
    for row, data in enumerate(sample_data_zh, 3):
        for col, value in enumerate(data, 1):
            ws_zh.cell(row=row, column=col, value=value)
    
    # Create FAQ format template (for Chinese users who might have single column data)
    ws_faq = wb.create_sheet("FAQ_Format")
    
    # FAQ headers - simple format that matches your uploaded file
    headers_faq = ['SKU', '内容', '图片链接']
    
    for col, header in enumerate(headers_faq, 1):
        cell = ws_faq.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
    
    # FAQ descriptions
    descriptions_faq = [
        '产品编号（可留空，系统自动生成）',
        'FAQ内容 - 问题和答案（必填）',
        '相关图片链接（可选）'
    ]
    
    for col, desc in enumerate(descriptions_faq, 1):
        ws_faq.cell(row=2, column=col, value=desc)
    
    # FAQ sample data
    sample_data_faq = [
        ['', '问：如何安装方向盘？答：首先连接线束，然后用螺栓固定。', 'https://example.com/steering.jpg'],
        ['', '问：方向盘控制器不工作怎么办？答：检查方向盘后面的线路连接。', ''],
        ['', '问：如何焊接框架？答：使用MIG焊接并配备适当的安全设备。', 'https://example.com/welding.jpg']
    ]
    
    for row, data in enumerate(sample_data_faq, 3):
        for col, value in enumerate(data, 1):
            ws_faq.cell(row=row, column=col, value=value)
    
    # Auto-adjust column widths for all sheets
    for ws in [ws_en, ws_zh, ws_faq]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create instructions sheet
    instructions_ws = wb.create_sheet("Instructions")
    instructions = [
        ["Multilingual Excel Upload Instructions / 多语言Excel上传说明", ""],
        ["", ""],
        ["English Instructions:", ""],
        ["Required Columns:", "SKU, Question, Answer"],
        ["Optional Columns:", "Image_Link, Category, Keywords"],
        ["", ""],
        ["中文说明:", ""],
        ["必需列:", "SKU, 问题, 答案"],
        ["可选列:", "图片链接, 类别, 关键词"],
        ["", ""],
        ["Supported Column Names / 支持的列名:", ""],
        ["SKU:", "SKU, sku, 产品编号, 产品代码, 编号, product_id"],
        ["Question:", "Question, 问题, 内容, Q, query, 疑问"],
        ["Answer:", "Answer, 答案, 回答, A, response, 解答"],
        ["Image_Link:", "Image_Link, 图片链接, 图片, image, img_link"],
        ["Category:", "Category, 类别, 分类, type, 种类"],
        ["Keywords:", "Keywords, 关键词, 关键字, tags, 标签"],
        ["", ""],
        ["File Formats / 文件格式:", ""],
        ["Supported:", ".xlsx, .xls"],
        ["Max Size:", "10MB"],
        ["", ""],
        ["Important Notes / 重要说明:", ""],
        ["• Column names are case-insensitive / 列名不区分大小写", ""],
        ["• You can use spaces or underscores / 可以使用空格或下划线", ""],
        ["• Empty rows will be skipped / 空行将被跳过", ""],
        ["• If SKU is missing, auto-generated IDs will be used / 如果缺少SKU，将自动生成ID", ""],
        ["• For FAQ format: put Q&A in one column / FAQ格式：问答放在一列中", ""],
        ["", ""],
        ["Template Sheets / 模板工作表:", ""],
        ["English_Template:", "Standard English format"],
        ["Chinese_Template:", "Standard Chinese format / 标准中文格式"],
        ["FAQ_Format:", "Simplified FAQ format / 简化FAQ格式"],
    ]
    
    for row, (key, value) in enumerate(instructions, 1):
        instructions_ws.cell(row=row, column=1, value=key)
        instructions_ws.cell(row=row, column=2, value=value)
        
        # Format headers
        if key and not value and row > 2:
            cell = instructions_ws.cell(row=row, column=1)
            cell.font = Font(bold=True)
    
    # Auto-adjust column widths for instructions
    for column in instructions_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 80)
        instructions_ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

def log_user_activity(user_ip: str, action: str, details: str = ""):
    """Log user activity for analytics"""
    logger.info(f"User Activity - IP: {user_ip}, Action: {action}, Details: {details}")

class SearchQueryProcessor:
    """Enhanced query processor optimized for semantic search ONLY"""
    
    def __init__(self):
        # Semantic mappings for better understanding
        self.semantic_mappings = {
            # Common technical terms
            'engine': ['发动机', '引擎', '马达', 'motor'],
            'transmission': ['变速箱', '传动', 'gearbox'],
            'brake': ['刹车', '制动', '刹车片'],
            'clutch': ['离合器', '离合'],
            'hydraulic': ['液压', '油压'],
            'electric': ['电动', '电气', '电力'],
            'fuel': ['燃料', '燃油', '汽油', 'gas', 'gasoline', 'diesel'],
            
            # Problem types
            'problem': ['问题', '故障', '毛病', 'issue', 'trouble'],
            'error': ['错误', '出错', 'fault'],
            'broken': ['坏了', '损坏', '破损', 'damaged'],
            'not working': ['不工作', '不运行', '失效'],
            
            # Installation terms
            'install': ['安装', '装配', '组装', 'mount', 'setup'],
            'replace': ['更换', '替换', '换', 'change'],
            'repair': ['修理', '维修', '修复', 'fix'],
            'maintenance': ['保养', '维护', '维修'],
            
            # Common products
            'filter': ['滤器', '过滤器', 'strainer'],
            'seal': ['密封', '密封件', 'gasket'],
            'bearing': ['轴承', '轴套'],
            'valve': ['阀门', '阀', 'valve'],
            'pump': ['泵', '水泵', '油泵'],
            'sensor': ['传感器', '感应器'],
            
            # Actions and Procedures
            'check': ['检查', '查看', 'inspect'],
            'test': ['测试', '检测', 'examine'],
            'adjust': ['调整', '调节', 'tune'],
            'clean': ['清洁', '清理', 'wash'],
            'lubricate': ['润滑', '加油', 'oil'],
            'tighten': ['拧紧', '固定', 'secure'],
            'loosen': ['松开', '放松', 'release'],
            
            # Common questions
            'how': ['如何', '怎么', '怎样'],
            'what': ['什么', '哪个'],
            'where': ['哪里', '在哪'],
            'when': ['什么时候', '何时'],
            'why': ['为什么', '为何'],
            'which': ['哪个', '哪一个'],
        }
        
        # Common stop words for filtering
        self.chinese_stops = {
            '的', '了', '在', '是', '我', '有', '和', '就', '不', '人', '都', '一', 
            '这', '那', '个', '上', '也', '很', '到', '说', '要', '去', '你', '会', 
            '着', '没有', '看', '好', '自己', '然后', '现在', '已经', '可以', '应该',
            '需要', '必须', '能够', '或者', '还是', '如果', '虽然', '因为', '所以',
            '但是', '问', '答'
        }
        
        self.english_stops = {
            'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 
            'of', 'with', 'by', 'from', 'as', 'is', 'are', 'was', 'were', 'be', 
            'been', 'being', 'have', 'has', 'had', 'do', 'does', 'did', 'will', 
            'would', 'could', 'should', 'can', 'may', 'might', 'must', 'this', 
            'that', 'these', 'those', 'i', 'you', 'he', 'she', 'it', 'we', 'they'
        }
    
    def clean_query(self, query: str) -> str:
        """Clean and normalize query while preserving ALL content including potential SKUs"""
        # Remove extra whitespace
        query = re.sub(r'\s+', ' ', query).strip()
        
        # Remove special characters but keep basic punctuation, Chinese characters, AND alphanumeric
        # This preserves SKU-like patterns as regular search terms
        query = re.sub(r'[^\w\s\-_.,!?，。！？\u4e00-\u9fff]', '', query)
        
        return query
    
    def extract_semantic_terms(self, query: str) -> List[str]:
        """Extract semantic terms WITHOUT removing SKU patterns - treat everything as semantic"""
        cleaned_query = query.strip()
        
        # NO SKU REMOVAL - treat potential SKUs as regular search terms
        # This is the key change - we don't filter out anything that looks like a SKU
        
        # Extract both Chinese and English terms, including alphanumeric patterns
        chinese_chars = re.findall(r'[\u4e00-\u9fff]+', cleaned_query)
        english_words = re.findall(r'[a-zA-Z]{1,}', cleaned_query.lower())  # Allow single chars
        alphanumeric = re.findall(r'[a-zA-Z]*\d+[a-zA-Z]*|\d*[a-zA-Z]+\d*', cleaned_query)  # Include alphanumeric patterns
        
        # Filter terms but keep more content
        chinese_terms = [term for term in chinese_chars 
                        if len(term) >= 1 and term not in self.chinese_stops]
        english_terms = [term for term in english_words 
                        if len(term) >= 1 and term not in self.english_stops]  # Reduced minimum length
        alphanumeric_terms = [term.lower() for term in alphanumeric if len(term) >= 1]
        
        # Start with original terms INCLUDING potential SKUs
        expanded_terms = chinese_terms + english_terms + alphanumeric_terms
        
        # Expand using semantic mappings
        for english_term in english_terms:
            if english_term in self.semantic_mappings:
                expanded_terms.extend(self.semantic_mappings[english_term])
        
        # Reverse lookup for Chinese terms
        for chinese_term in chinese_terms:
            for english_key, translations in self.semantic_mappings.items():
                if chinese_term in translations:
                    expanded_terms.append(english_key)
                    # Add other translations too
                    expanded_terms.extend([t for t in translations if t != chinese_term])
        
        # Remove duplicates while preserving order
        seen = set()
        unique_terms = []
        for term in expanded_terms:
            if term not in seen:
                seen.add(term)
                unique_terms.append(term)
        
        # Log the expansion for debugging
        logger.info(f"🔍 Semantic terms extracted: {unique_terms[:10]}...")  # Show first 10
        
        return unique_terms[:25]  # Increased limit for better semantic context
    
    def get_query_intent(self, query: str) -> str:
        """Detect query intent for better RAG prompting"""
        query_lower = query.lower()
        
        if any(term in query_lower for term in ['how', 'install', '安装', '如何', 'setup']):
            return 'installation'
        elif any(term in query_lower for term in ['problem', 'error', '问题', '故障', 'troubleshoot', 'fix']):
            return 'troubleshooting'
        elif any(term in query_lower for term in ['what', 'which', '什么', '哪个', 'specifications', 'spec']):
            return 'information'
        elif any(term in query_lower for term in ['replace', 'repair', '更换', '修理', 'maintenance']):
            return 'maintenance'
        else:
            return 'general'